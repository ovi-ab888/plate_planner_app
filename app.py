import streamlit as st
import pandas as pd
from io import BytesIO
import pypdf
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="PDF to Excel Converter", page_icon="📄", layout="wide")

st.markdown("""
<style>
    .stApp { background-color: #0e1117; color: white; }
    .main-title {
        background: linear-gradient(135deg, #11998e 0%, #38ef7d 100%);
        padding: 1.5rem;
        border-radius: 10px;
        text-align: center;
        color: white;
        margin-bottom: 2rem;
    }
    .success-box {
        background-color: #1e3a1e;
        padding: 15px;
        border-radius: 8px;
        border-left: 5px solid #2e7d32;
    }
</style>
<div class="main-title">
    <h1>📄 WorkOrder PDF to Excel (XLSX) Converter</h1>
    <p>Upload Work Order PDF and Download Clean, Styled Excel File</p>
</div>
""", unsafe_allow_html=True)

uploaded_file = st.file_uploader("আপনার WorkOrder PDF ফাইলটি এখানে আপলোড করুন", type=["pdf"])

if uploaded_file is not None:
    try:
        # ১. পিডিএফ থেকে টেক্সট এক্সট্রাক্ট করা
        reader = pypdf.PdfReader(uploaded_file)
        full_text = ""
        for page in reader.pages:
            full_text += page.extract_text() + "\n"
            
        lines = full_text.split('\n')
        raw_data = []
        
        # ২. টেক্সট লাইনগুলোকে রো (Row) আকারে স্যানিটাইজ ও ফিল্টার করা
        for line in lines:
            line = line.strip()
            
            # হেডার বা অপ্রয়োজনীয় লাইন বাদ দেওয়া
            if "SO NO." in line or "PRODUCT CODE" in line or not line:
                continue
                
            # কমা বা কোটেশন যুক্ত ডাটা রো আইডেন্টিফাই করা
            if '",' in line or ',"' in line or (line.startswith('"') and line.endswith('"')):
                # লাইনের ভেতরের অনাকাঙ্ক্ষিত \n রিমুভ করা
                sanitized_line = line.replace('\n', ' ')
                
                # CSV রিডার দিয়ে নিখুঁত কলাম পার্সিং
                csv_reader = csv.reader([sanitized_line])
                parts = next(csv_reader)
                parts = [p.strip() for p in parts]
                
                # সঠিক টেবিল রো কিনা তা নিশ্চিত করা (SO Number সংখ্যা হতে হবে)
                if len(parts) >= 6 and parts[0].replace(' ', '').isdigit():
                    so_no = parts[0]
                    prod_desc = parts[1]
                    barcode_size = parts[2]
                    price = parts[3]
                    story = parts[4]
                    qty = parts[5]
                    
                    raw_data.append([so_no, prod_desc, barcode_size, price, story, float(qty)])
        
        if raw_data:
            # ৩. ওপেনপিক্সেল (openpyxl) দিয়ে এক্সেল ডিজাইন ও স্টাইলিং
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "WorkOrder Quantities"
            ws.views.sheetView[0].showGridLines = True
            
            # স্টাইল গাইডলাইনস
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # ক্লাসিক ব্লু হেডার
            header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            data_font = Font(name="Segoe UI", size=10)
            total_font = Font(name="Segoe UI", size=11, bold=True)
            total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            
            thin_border = Border(
                left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
                top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
            )
            double_bottom_border = Border(
                left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
                top=Side(style='thin', color='D9D9D9'), bottom=Side(style='double', color='1F4E78')
            )
            
            # হেডার পুশ করা
            headers = ["SL", "SO No.", "Product Code & Description", "Barcode & Size", "Price", "Story Name", "Quantity"]
            ws.append(headers)
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # ডাটা রো পুশ করা
            for idx, row in enumerate(raw_data, 1):
                row_data = [idx, row[0], row[1], row[2], row[3], row[4], row[5]]
                ws.append(row_data)
                row_num = idx + 1
                
                for col_num in range(1, len(headers) + 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.font = data_font
                    cell.border = thin_border
                    
                    # কলাম অ্যালাইনমেন্ট
                    if col_num in [1, 2, 5, 6]:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    elif col_num in [3, 4]:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    elif col_num == 7:
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        cell.number_format = '#,##0.00' # নাম্বার ফরম্যাটিং
            
            # ৪. টোটাল রো এবং এক্সেল SUM ফর্মুলা যোগ করা
            last_row = len(raw_data) + 2
            ws.cell(row=last_row, column=1, value="📊").alignment = Alignment(horizontal="center")
            ws.cell(row=last_row, column=2, value="TOTAL").font = total_font
            ws.cell(row=last_row, column=2).alignment = Alignment(horizontal="center")
            
            for col_num in range(1, len(headers)):
                cell = ws.cell(row=last_row, column=col_num)
                cell.fill = total_fill
                cell.font = total_font
                cell.border = double_bottom_border
                
            total_qty_cell = ws.cell(row=last_row, column=7, value=f"=SUM(G2:G{last_row-1})")
            total_qty_cell.fill = total_fill
            total_qty_cell.font = total_font
            total_qty_cell.alignment = Alignment(horizontal="right", vertical="center")
            total_qty_cell.number_format = '#,##0.00'
            total_qty_cell.border = double_bottom_border
            
            # কলামের উইড্থ (Width) অটো-অ্যাডজাস্ট করা
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max(max_len + 3, 11)
                
            ws.row_dimensions[1].height = 28
            for r in range(2, last_row + 1):
                ws.row_dimensions[r].height = 22
                
            # মেমোরি বাফারে ফাইলটি সেভ করা
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            st.markdown("""
            <div class="success-box">
                <h3 style='color: #4caf50; margin: 0;'>✅ কনভার্সন সফল হয়েছে!</h3>
                <p style='margin: 5px 0 0 0;'>পিডিএফ থেকে সবগুলো আইটেম এবং Quantity কলাম সুন্দরভাবে এক্সেল শিটে সাজানো হয়েছে। নিচের বাটনে ক্লিক করে ডাউনলোড করুন।</p>
            </div>
            """, unsafe_allow_html=True)
            st.markdown("<br>", unsafe_allow_html=True)
            
            st.download_button(
                label="📥 Download Clean Excel (XLSX)",
                data=excel_buffer,
                file_name="Converted_WorkOrder_Data.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        else:
            st.error("❌ পিডিএফ থেকে কোনো ডাটা রো রিড করা যায়নি। ফাইল ফরম্যাটটি চেক করুন।")
            
    except Exception as e:
        st.error(f"❌ কনভার্ট করতে সমস্যা হয়েছে: {str(e)}")
